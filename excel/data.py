import openpyxl

def get_data():
    workbook = openpyxl.load_workbook("..//excel//appiumdata.xlsx")
    sheet_name = "vitok"
    sheet = workbook[sheet_name]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    
    mainList = []
    for i in range(2, totalrows +1):
        dataList = []
        for j in range(1, totalcols +1):
            data = sheet.cell(row=i,column =j).value
            dataList.append(j,data)
        mainList.append(i,dataList)
    return mainList

def login_data():
    # data = get_data()
    # valid_data = data[0]
    # invalid_data = data[1]
    return {
        # "valid_phone": valid_data[0],
        # "valid_otp": valid_data[1],
        # "invalid_phone": invalid_data[0],
        # "invalid_otp": invalid_data[1],
        "valid_phone": "9876543210",
        "valid_otp": "0",
        "DOB_date":"8",
        "DOB_month":"March",
        "DOB_year":"2001",
        "nick_name": "Testy",
        "referral_code": "ZEZWLFW6"
        
    }
    